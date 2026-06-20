#!/usr/bin/env python3
"""
Summarize alarm identifiers from one or more Excel .xlsx log files.

Usage:
  python3 multi_alarm_excel_summary.py CN-ALEAF-1-260616-logfile.xlsx SLEAF_POD2_CN-260616-logfile.xlsx

If no files are provided, the script prompts for file names. Separate multiple
file names with spaces or commas.

Output:
  An XLSX summary table whose file name includes the generation date/time by default.
  Use --csv to generate CSV instead.
  Each alarm identifier has one row, with one log-content example.
  The configured ignored alarm identifiers are placed at the end and marked
  as "无需关注".
"""

from __future__ import annotations

import argparse
import csv
import shlex
import sys
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional
from zipfile import ZipFile

import openpyxl
from openpyxl.styles import Alignment, Font

__version__ = "1.0.0"


IGNORED_ALERT_IDS = {
    "SHELL/6/SHELL_CMD",
    "PING/6/PING_STATISTICS",
    "SSHS/6/SSHS_LOG",
    "SSHS/6/SSHS_DISCONNECT",
    "SSHS/6/SSHS_CONNECT",
    "SHELL/5/SHELL_LOGIN",
    "SHELL/5/SHELL_LOGOUT",
    "NETCONF/6/SSH_XML_LOGOUT",
    "NETCONF/6/SSH_XML_LOGIN",
}

SPREADSHEET_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
RELATIONSHIP_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

NEW_FORMAT_COLUMNS = {
    "sequence": "日志序号",
    "date": "日期",
    "time": "时间",
    "alert_id": "模块名称/等级/助记符",
    "detail": "日志内容",
}

OLD_FORMAT_COLUMNS = {
    "date": "A",
    "time": "B",
    "alert_id": "F",
    "detail": "G",
}

FIXED_XLSX_COLUMN_WIDTHS = {
    "A": 51.76,
    "B": 9.93,
    "C": 10.31,
    "D": 31,
    "E": 14,
    "F": 21,
    "G": 80,
}


def column_letters(cell_ref: str) -> str:
    return "".join(ch for ch in cell_ref if ch.isalpha())


def load_shared_strings(workbook_zip: ZipFile) -> list[str]:
    try:
        shared_strings_file = workbook_zip.open("xl/sharedStrings.xml")
    except KeyError:
        return []

    strings: list[str] = []
    for _event, elem in ET.iterparse(shared_strings_file, events=("end",)):
        if elem.tag == SPREADSHEET_NS + "si":
            strings.append("".join((text.text or "") for text in elem.iter(SPREADSHEET_NS + "t")))
            elem.clear()
    return strings


def cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")

    if cell_type == "inlineStr":
        return "".join((text.text or "") for text in cell.iter(SPREADSHEET_NS + "t"))

    value = cell.find(SPREADSHEET_NS + "v")
    if value is None:
        return ""

    raw_value = value.text or ""
    if cell_type == "s":
        try:
            return shared_strings[int(raw_value)]
        except (ValueError, IndexError):
            return raw_value

    return raw_value


def normalize_text(value: str) -> str:
    return value.strip().replace("\r\n", " ").replace("\n", " ")


def excel_datetime_text(date_value: str, time_value: str) -> str:
    try:
        date_number = float(date_value)
        time_number = float(time_value or "0")
    except ValueError:
        if date_value and time_value:
            return f"{date_value} {time_value}"
        return date_value or time_value

    dt = datetime(1899, 12, 30) + timedelta(days=date_number + time_number)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def workbook_sheets(workbook_zip: ZipFile) -> list[tuple[str, str]]:
    workbook = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
    relationships = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
    rel_id_to_target = {}

    for rel in relationships:
        target = rel.attrib["Target"].lstrip("/")
        if not target.startswith("xl/"):
            target = "xl/" + target
        rel_id_to_target[rel.attrib["Id"]] = target

    sheets: list[tuple[str, str]] = []
    for sheet in workbook.find(SPREADSHEET_NS + "sheets"):
        sheet_name = sheet.attrib["name"]
        relationship_id = sheet.attrib[RELATIONSHIP_NS + "id"]
        sheets.append((sheet_name, rel_id_to_target[relationship_id]))

    return sheets


def row_values_by_column(row: ET.Element, shared_strings: list[str]) -> dict[str, str]:
    values = {}
    for cell in row.findall(SPREADSHEET_NS + "c"):
        col = column_letters(cell.attrib.get("r", ""))
        values[col] = normalize_text(cell_text(cell, shared_strings))
    return values


def detect_columns(header_values: dict[str, str]) -> Optional[dict[str, str]]:
    header_to_col = {value: col for col, value in header_values.items() if value}
    new_columns = {}

    for key, header in NEW_FORMAT_COLUMNS.items():
        col = header_to_col.get(header)
        if col:
            new_columns[key] = col

    if "alert_id" in new_columns and "detail" in new_columns:
        return new_columns

    # Backward-compatible fallback for the old exported alarm tables.
    return dict(OLD_FORMAT_COLUMNS)


def parse_alarm_file(input_file: Path, aggregate: dict[str, dict]) -> int:
    total_rows = 0

    with ZipFile(input_file) as workbook_zip:
        shared_strings = load_shared_strings(workbook_zip)

        for sheet_name, sheet_path in workbook_sheets(workbook_zip):
            columns: Optional[dict[str, str]] = None

            with workbook_zip.open(sheet_path) as sheet_file:
                for _event, row in ET.iterparse(sheet_file, events=("end",)):
                    if row.tag != SPREADSHEET_NS + "row":
                        continue

                    values = row_values_by_column(row, shared_strings)

                    if columns is None:
                        columns = detect_columns(values)
                        row.clear()
                        continue

                    alert_id = values.get(columns.get("alert_id", ""), "")
                    detail = values.get(columns.get("detail", ""), "")
                    date_value = values.get(columns.get("date", ""), "")
                    time_value = values.get(columns.get("time", ""), "")
                    sequence = values.get(columns.get("sequence", ""), "")

                    if alert_id:
                        total_rows += 1

                        record = aggregate[alert_id]
                        record["total"] += 1
                        record["files"][input_file.name] += 1

                        if not record["example_detail"]:
                            record["example_detail"] = detail
                            record["example_file"] = input_file.name
                            record["example_sequence"] = sequence
                            record["example_datetime"] = excel_datetime_text(date_value, time_value)

                    row.clear()

    return total_rows


def output_base_name() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def summary_file_label(path: Path) -> str:
    name = path.name
    for suffix in ("-logfile.xlsx", "-logfile.csv", ".xlsx", ".csv"):
        if name.endswith(suffix):
            name = name[: -len(suffix)]
            break

    parts = name.split("-")
    if parts and parts[-1].isdigit() and len(parts[-1]) == 6:
        parts = parts[:-1]
    if parts:
        return "-".join(parts)
    return name


def append_rows(sheet, rows: list[list[object]]) -> None:
    for row in rows:
        sheet.append(row)


def write_summary(
    output_file: Path,
    aggregate: dict[str, dict],
    input_files: list[Path],
    file_total_rows: dict[str, int],
) -> None:
    file_names = [path.name for path in input_files]
    file_labels = [summary_file_label(path) for path in input_files]
    headers = [
        "告警标识符",
        "总次数",
        "备注",
        "示例来源文件",
        "示例日志序号",
        "示例时间",
        "示例告警内容",
    ]
    headers.extend(file_labels)

    def sort_key(item: tuple[str, dict]) -> tuple[int, int, str]:
        alert_id, data = item
        ignored_rank = 1 if alert_id in IGNORED_ALERT_IDS else 0
        return (ignored_rank, -data["total"], alert_id)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "告警标识符汇总"

    rows = [headers]
    for alert_id, data in sorted(aggregate.items(), key=sort_key):
        remark = "无需关注" if alert_id in IGNORED_ALERT_IDS else ""
        row = [
            alert_id,
            data["total"],
            remark,
            data["example_file"],
            data["example_sequence"],
            data["example_datetime"],
            data["example_detail"],
        ]
        row.extend(data["files"].get(name, 0) for name in file_names)
        rows.append(row)

    rows.extend(
        [
            [],
            ["数据来源文件汇总"],
            ["来源文件数", len(file_names)],
            ["来源文件总告警行数", sum(file_total_rows.values())],
            [],
            ["文件名", "总告警数"],
        ]
    )
    for name in file_names:
        rows.append([name, file_total_rows.get(name, 0)])

    append_rows(sheet, rows)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")

    for row in sheet.iter_rows():
        for cell in row:
            should_center = cell.column == 2 or cell.column == 3 or cell.column == 5 or cell.column >= 8
            cell.alignment = Alignment(
                horizontal="center" if should_center else None,
                wrap_text=False,
                vertical="top",
            )

    if sheet.max_row > 1:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{sheet.cell(row=1, column=sheet.max_column).column_letter}{len(aggregate) + 1}"

    for column, width in FIXED_XLSX_COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width
    for column_index in range(8, sheet.max_column + 1):
        column_letter = sheet.cell(row=1, column=column_index).column_letter
        sheet.column_dimensions[column_letter].width = 14.3
    workbook.save(output_file)


def write_summary_csv(
    output_file: Path,
    aggregate: dict[str, dict],
    input_files: list[Path],
    file_total_rows: dict[str, int],
) -> None:
    file_names = [path.name for path in input_files]
    file_labels = [summary_file_label(path) for path in input_files]
    headers = [
        "告警标识符",
        "总次数",
        "备注",
        "示例来源文件",
        "示例日志序号",
        "示例时间",
        "示例告警内容",
    ]
    headers.extend(file_labels)

    def sort_key(item: tuple[str, dict]) -> tuple[int, int, str]:
        alert_id, data = item
        ignored_rank = 1 if alert_id in IGNORED_ALERT_IDS else 0
        return (ignored_rank, -data["total"], alert_id)

    with output_file.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(headers)

        for alert_id, data in sorted(aggregate.items(), key=sort_key):
            remark = "无需关注" if alert_id in IGNORED_ALERT_IDS else ""
            row = [
                alert_id,
                data["total"],
                remark,
                data["example_file"],
                data["example_sequence"],
                data["example_datetime"],
                data["example_detail"],
            ]
            row.extend(data["files"].get(name, 0) for name in file_names)
            writer.writerow(row)

        writer.writerow([])
        writer.writerow(["数据来源文件汇总"])
        writer.writerow(["来源文件数", len(file_names)])
        writer.writerow(["来源文件总告警行数", sum(file_total_rows.values())])
        writer.writerow([])
        writer.writerow(["文件名", "该文件总告警行数"])
        for name in file_names:
            writer.writerow([name, file_total_rows.get(name, 0)])


def split_prompted_files(raw_input: str) -> list[str]:
    normalized = raw_input.replace(",", " ")
    return shlex.split(normalized)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="汇总一个或多个 .xlsx 日志表格的告警标识符次数，并提供每类告警示例。"
    )
    parser.add_argument("--version", action="version", version=f"%(prog)s {__version__}")
    parser.add_argument("input_files", nargs="*", help="待处理的 .xlsx 文件，最多建议 20 个")
    parser.add_argument(
        "-o",
        "--output-dir",
        default="alarm_summary_output",
        help="输出目录，默认 alarm_summary_output",
    )
    parser.add_argument(
        "--csv",
        action="store_true",
        help="输出 CSV 文件；默认输出带固定列宽的 Excel 文件",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_names = args.input_files

    if not input_names:
        raw = input("请输入要处理的 Excel 文件名，多个文件用空格或逗号分隔：").strip()
        input_names = split_prompted_files(raw)

    if not input_names:
        print("未提供输入文件。", file=sys.stderr)
        return 1

    if len(input_names) > 20:
        print("输入文件超过 20 个，请分批处理。", file=sys.stderr)
        return 1

    input_files = [Path(name).expanduser() for name in input_names]
    missing = [str(path) for path in input_files if not path.exists()]
    if missing:
        print("以下文件不存在：", file=sys.stderr)
        for path in missing:
            print(f"  {path}", file=sys.stderr)
        return 1

    invalid = [str(path) for path in input_files if path.suffix.lower() != ".xlsx"]
    if invalid:
        print("当前脚本只支持 .xlsx 文件：", file=sys.stderr)
        for path in invalid:
            print(f"  {path}", file=sys.stderr)
        return 1

    aggregate = defaultdict(
        lambda: {
            "total": 0,
            "files": Counter(),
            "example_detail": "",
            "example_file": "",
            "example_sequence": "",
            "example_datetime": "",
        }
    )
    file_total_rows: dict[str, int] = {}

    for index, input_file in enumerate(input_files, start=1):
        print(f"[{index}/{len(input_files)}] 正在处理：{input_file}")
        total_rows = parse_alarm_file(input_file, aggregate)
        file_total_rows[input_file.name] = total_rows
        print(f"    完成，告警行数：{total_rows}")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(exist_ok=True)
    base_name = output_base_name()
    if args.csv:
        output_file = output_dir / f"{base_name}_告警标识符汇总_含示例.csv"
        write_summary_csv(output_file, aggregate, input_files, file_total_rows)
    else:
        output_file = output_dir / f"{base_name}_告警标识符汇总_含示例.xlsx"
        write_summary(output_file, aggregate, input_files, file_total_rows)

    ignored_total = sum(data["total"] for alert_id, data in aggregate.items() if alert_id in IGNORED_ALERT_IDS)
    print("处理完成")
    print(f"输入文件数：{len(input_files)}")
    print(f"总告警行数：{sum(file_total_rows.values())}")
    print(f"告警标识符种类数：{len(aggregate)}")
    print(f"无需关注告警总行数：{ignored_total}")
    print(f"输出文件：{output_file.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
