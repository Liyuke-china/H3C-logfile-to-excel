#!/usr/bin/env python3
"""
Convert raw device logfile archives to Excel and generate alarm summaries in one run.
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("缺少依赖 openpyxl，请先执行：python3 -m pip install openpyxl") from exc

import log_to_excel
import multi_alarm_excel_summary as alarm_summary

__version__ = "1.1.0"


def log(message: str) -> None:
    print(message, flush=True)


def output_base_name() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def detail_output_path(input_path: Path, output_name: Optional[str] = None) -> Path:
    if output_name:
        return log_to_excel.output_directory() / Path(output_name).name
    return log_to_excel.default_output_path(input_path)


def aggregate_from_log_excel(input_file: Path, aggregate: dict[str, dict]) -> int:
    workbook = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    if "logs" not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"{input_file} 中没有 logs 工作表")

    sheet = workbook["logs"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    column_map = {name: index for index, name in enumerate(header)}

    required_columns = ["日志序号", "日期", "时间", "模块名称/等级/助记符", "日志内容"]
    missing = [name for name in required_columns if name not in column_map]
    if missing:
        workbook.close()
        raise ValueError(f"{input_file} 缺少必要列：{', '.join(missing)}")

    total_rows = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        alert_id = row[column_map["模块名称/等级/助记符"]]
        if not alert_id:
            continue

        total_rows += 1
        record = aggregate[str(alert_id)]
        record["total"] += 1
        record["files"][input_file.name] += 1

        if not record["example_detail"]:
            date_value = row[column_map["日期"]]
            time_value = row[column_map["时间"]]
            example_time = ""
            if date_value and time_value:
                example_time = f"{date_value:%Y-%m-%d} {time_value:%H:%M:%S}"
            elif date_value:
                example_time = str(date_value)
            elif time_value:
                example_time = str(time_value)

            record["example_detail"] = row[column_map["日志内容"]] or ""
            record["example_file"] = input_file.name
            record["example_sequence"] = row[column_map["日志序号"]] or ""
            record["example_datetime"] = example_time

    workbook.close()
    return total_rows


def write_summary_outputs(
    aggregate: dict[str, dict],
    detail_files: list[Path],
    file_total_rows: dict[str, int],
    output_dir: Path,
    include_csv: bool,
) -> tuple[Path, Optional[Path]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    base_name = output_base_name()

    summary_xlsx = output_dir / f"{base_name}_告警标识符汇总_含示例.xlsx"
    alarm_summary.write_summary(summary_xlsx, aggregate, detail_files, file_total_rows)

    summary_csv = None
    if include_csv:
        summary_csv = output_dir / f"{base_name}_告警标识符汇总_含示例.csv"
        alarm_summary.write_summary_csv(summary_csv, aggregate, detail_files, file_total_rows)

    return summary_xlsx, summary_csv


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="一次性完成原始日志包转 Excel，并生成告警标识符汇总分析。"
    )
    parser.add_argument("--version", action="version", version=f"%(prog)s {__version__}")
    parser.add_argument("input_files", nargs="+", help="输入文件，支持 .tar.gz/.tgz、.log.gz 或 .log")
    parser.add_argument(
        "-o",
        "--output",
        help="仅处理单个输入文件时可指定日志明细 Excel 文件名，固定保存到 output 文件夹",
    )
    parser.add_argument(
        "--summary-output-dir",
        default="alarm_summary_output",
        help="汇总分析输出目录，默认 alarm_summary_output",
    )
    parser.add_argument(
        "--csv",
        action="store_true",
        help="除默认汇总 Excel 外，额外生成一份汇总 CSV",
    )
    parser.add_argument("--open", action="store_true", help="完成后在 macOS 中打开汇总 Excel")
    return parser


def main(argv: Optional[list[str]] = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    input_paths = [Path(name).expanduser().resolve() for name in args.input_files]
    missing = [str(path) for path in input_paths if not path.exists()]
    if missing:
        print("以下文件不存在：", file=sys.stderr)
        for path in missing:
            print(f"  {path}", file=sys.stderr)
        return 1

    if args.output and len(input_paths) != 1:
        parser.error("-o/--output 只能在处理单个输入文件时使用")

    aggregate = defaultdict(
        lambda: {
            "total": 0,
            "files": Counter(),
            "example_detail": "",
            "example_file": "",
            "example_sequence": "",
            "example_datetime": "",
        }
    )
    detail_files: list[Path] = []
    file_total_rows: dict[str, int] = {}

    for index, input_path in enumerate(input_paths, start=1):
        log(f"[{index}/{len(input_paths)}] 生成日志明细：{input_path}")
        output_path = detail_output_path(input_path, args.output)
        result = log_to_excel.parse_logfile_to_excel(input_path, output_path)
        detail_files.append(output_path)
        log(
            "    明细完成："
            f"日志记录 {result['log_records']:,}，"
            f"续行 {result['continuation_lines']:,}，"
            f"无法归属 {result['unparsed_lines']:,}"
        )

        log(f"    汇总统计：{output_path}")
        total_rows = aggregate_from_log_excel(output_path, aggregate)
        file_total_rows[output_path.name] = total_rows
        log(f"    告警行数：{total_rows:,}")

    summary_xlsx, summary_csv = write_summary_outputs(
        aggregate=aggregate,
        detail_files=detail_files,
        file_total_rows=file_total_rows,
        output_dir=Path(args.summary_output_dir),
        include_csv=args.csv,
    )

    ignored_total = sum(
        data["total"] for alert_id, data in aggregate.items() if alert_id in alarm_summary.IGNORED_ALERT_IDS
    )
    log("处理完成")
    log(f"输入文件数：{len(input_paths)}")
    log(f"总告警行数：{sum(file_total_rows.values()):,}")
    log(f"告警标识符种类数：{len(aggregate)}")
    log(f"无需关注告警总行数：{ignored_total:,}")
    log(f"日志明细文件：{', '.join(str(path) for path in detail_files)}")
    log(f"汇总 Excel：{summary_xlsx.resolve()}")
    if summary_csv:
        log(f"汇总 CSV：{summary_csv.resolve()}")

    if args.open:
        import subprocess

        subprocess.run(["open", str(summary_xlsx)], check=False)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
