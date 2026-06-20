#!/usr/bin/env python3
import argparse
import gzip
import io
import re
import subprocess
import tarfile
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterator, Optional, TextIO

__version__ = "1.0.0"

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, NamedStyle
except ImportError as exc:
    raise SystemExit("缺少依赖 openpyxl，请先执行：python3 -m pip install openpyxl") from exc


LOG_PATTERN = re.compile(
    r"^%@(?P<sequence>\d+)%"
    r"(?P<timestamp>[A-Za-z]{3}\s+\d+\s+\d{2}:\d{2}:\d{2}:\d{3}\s+\d{4})"
    r"\s+(?P<device>\S+)"
    r"\s+(?P<key>[^\s:]+):\s?"
    r"(?P<content>.*)$"
)

MONTH_MAP = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12,
}

LOG_HEADERS = [
    "日志文件",
    "原始行号",
    "日志序号",
    "日期",
    "时间",
    "设备名",
    "模块名称",
    "等级",
    "助记符",
    "模块名称/等级/助记符",
    "日志内容",
]

UNPARSED_HEADERS = ["日志文件", "原始行号", "原始内容", "原因"]


@dataclass
class LogRecord:
    source_log_file: str
    line_number: int
    sequence: str
    date: datetime
    time: object
    device: str
    module: str
    level: str
    mnemonic: str
    key: str
    content: str


def log(message: str) -> None:
    print(message, flush=True)


def remove_illegal_characters(value):
    if isinstance(value, str):
        # Excel allows line breaks in cells, so keep \n and \r for multiline logs.
        return re.sub(r"[\x00-\x09\x0B\x0C\x0E-\x1F\x7F]", "", value)
    return value


def parse_timestamp(timestamp: str) -> tuple[datetime, object]:
    parts = timestamp.split()
    if len(parts) != 4:
        raise ValueError(f"invalid timestamp: {timestamp}")

    month_name, day, time_text, year = parts
    if month_name not in MONTH_MAP:
        raise ValueError(f"invalid month: {month_name}")

    hour, minute, second, millisecond = time_text.split(":")
    parsed_date = datetime(int(year), MONTH_MAP[month_name], int(day))
    parsed_time = datetime.strptime(
        f"{hour}:{minute}:{second}.{millisecond}", "%H:%M:%S.%f"
    ).time()
    return parsed_date, parsed_time


def parse_log_line(line: str, source_log_file: str, line_number: int) -> LogRecord:
    match = LOG_PATTERN.match(line)
    if not match:
        raise ValueError("not a logfile record")

    key = match.group("key")
    key_parts = key.split("/")
    if len(key_parts) != 3:
        raise ValueError(f"invalid module/level/mnemonic: {key}")

    parsed_date, parsed_time = parse_timestamp(match.group("timestamp"))
    return LogRecord(
        source_log_file=source_log_file,
        line_number=line_number,
        sequence=match.group("sequence"),
        date=parsed_date,
        time=parsed_time,
        device=match.group("device"),
        module=key_parts[0],
        level=key_parts[1],
        mnemonic=key_parts[2],
        key=key,
        content=match.group("content"),
    )


def record_to_row(record: LogRecord) -> list:
    return [
        record.source_log_file,
        record.line_number,
        record.sequence,
        record.date,
        record.time,
        record.device,
        record.module,
        record.level,
        record.mnemonic,
        record.key,
        record.content,
    ]


def normalize_log_file_name(source_name: str) -> str:
    name = Path(source_name).name
    for suffix in (".log.gz", ".log"):
        if name.endswith(suffix):
            return name[: -len(suffix)]
    return Path(name).stem


def iter_plain_log(path: Path) -> Iterator[tuple[str, int, str]]:
    with path.open("r", encoding="utf-8", errors="replace", newline="") as file:
        source_log_file = normalize_log_file_name(path.name)
        yield from iter_text_lines(file, source_log_file)


def iter_gzip_log(path: Path) -> Iterator[tuple[str, int, str]]:
    with gzip.open(path, "rt", encoding="utf-8", errors="replace", newline="") as file:
        source_log_file = normalize_log_file_name(path.name)
        yield from iter_text_lines(file, source_log_file)


def iter_text_lines(file: TextIO, source_log_file: str) -> Iterator[tuple[str, int, str]]:
    for line_number, line in enumerate(file, start=1):
        yield source_log_file, line_number, line.rstrip("\r\n")


def iter_tar_logs(path: Path) -> Iterator[tuple[str, int, str]]:
    with tarfile.open(path, "r:gz") as tar:
        for member in tar.getmembers():
            if not member.isfile():
                continue
            if not (member.name.endswith(".log") or member.name.endswith(".log.gz")):
                continue

            extracted = tar.extractfile(member)
            if extracted is None:
                continue

            if member.name.endswith(".gz"):
                binary_file = gzip.GzipFile(fileobj=extracted)
            else:
                binary_file = extracted

            with io.TextIOWrapper(
                binary_file, encoding="utf-8", errors="replace", newline=""
            ) as text_file:
                source_log_file = normalize_log_file_name(member.name)
                yield from iter_text_lines(text_file, source_log_file)


def iter_input_lines(input_path: Path) -> Iterator[tuple[str, int, str]]:
    if input_path.name.endswith((".tar.gz", ".tgz")):
        yield from iter_tar_logs(input_path)
    elif input_path.name.endswith(".log.gz"):
        yield from iter_gzip_log(input_path)
    elif input_path.name.endswith(".log"):
        yield from iter_plain_log(input_path)
    else:
        raise ValueError("输入文件仅支持 .tar.gz/.tgz、.log.gz 或 .log")


def append_header(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")


def setup_workbook_styles(workbook: openpyxl.Workbook) -> None:
    if "date_style" not in workbook.named_styles:
        workbook.add_named_style(NamedStyle(name="date_style", number_format="YYYY-MM-DD"))
    if "time_style" not in workbook.named_styles:
        workbook.add_named_style(NamedStyle(name="time_style", number_format="HH:MM:SS.000"))


def set_logs_sheet_style(sheet) -> None:
    widths = {
        "A": 12,
        "B": 10,
        "C": 12,
        "D": 12,
        "E": 15,
        "F": 36,
        "G": 14,
        "H": 8,
        "I": 28,
        "J": 36,
        "K": 120,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:K{sheet.max_row}"

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        row[3].style = "date_style"
        row[4].style = "time_style"
        row[10].alignment = Alignment(wrap_text=True, vertical="top")


def set_summary_sheet_style(sheet) -> None:
    for column, width in {"A": 26, "B": 22, "C": 22, "D": 22}.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top")
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def set_unparsed_sheet_style(sheet) -> None:
    for column, width in {"A": 12, "B": 10, "C": 120, "D": 28}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def write_counter_section(sheet, title: str, counter: Counter, start_row: int) -> int:
    sheet.cell(row=start_row, column=1, value=title)
    sheet.cell(row=start_row, column=1).font = Font(bold=True)
    sheet.cell(row=start_row + 1, column=1, value="名称")
    sheet.cell(row=start_row + 1, column=2, value="次数")
    sheet.cell(row=start_row + 1, column=1).font = Font(bold=True)
    sheet.cell(row=start_row + 1, column=2).font = Font(bold=True)

    row = start_row + 2
    for name, count in counter.most_common():
        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=2, value=count)
        row += 1
    return row + 1


def output_directory() -> Path:
    return Path(__file__).resolve().parent / "output"


def default_output_path(input_path: Path) -> Path:
    output_dir = output_directory()
    name = input_path.name
    for suffix in (".tar.gz", ".tgz", ".log.gz", ".log"):
        if name.endswith(suffix):
            name = name[: -len(suffix)]
            break
    return output_dir / f"{name}.xlsx"


def requested_output_path(output_name: str) -> Path:
    return output_directory() / Path(output_name).name


def parse_logfile_to_excel(input_path: Path, output_path: Path) -> dict:
    workbook = openpyxl.Workbook()
    setup_workbook_styles(workbook)

    logs_sheet = workbook.active
    logs_sheet.title = "logs"
    summary_sheet = workbook.create_sheet("summary")
    unparsed_sheet = workbook.create_sheet("unparsed")

    append_header(logs_sheet, LOG_HEADERS)
    append_header(unparsed_sheet, UNPARSED_HEADERS)

    total_lines = 0
    log_records = 0
    continuation_lines = 0
    unparsed_lines = 0
    log_file_counter = Counter()
    module_counter = Counter()
    level_counter = Counter()
    mnemonic_counter = Counter()
    current_record: Optional[LogRecord] = None

    def flush_current_record() -> None:
        nonlocal current_record, log_records
        if current_record is None:
            return

        row = [remove_illegal_characters(value) for value in record_to_row(current_record)]
        logs_sheet.append(row)
        log_records += 1
        log_file_counter[current_record.source_log_file] += 1
        module_counter[current_record.module] += 1
        level_counter[current_record.level] += 1
        mnemonic_counter[current_record.mnemonic] += 1
        current_record = None

    for source_log_file, line_number, line in iter_input_lines(input_path):
        total_lines += 1
        if total_lines % 100000 == 0:
            log(f"已读取 {total_lines:,} 行，已解析 {log_records:,} 条日志")

        try:
            next_record = parse_log_line(line, source_log_file, line_number)
        except ValueError as exc:
            if current_record is not None:
                current_record.content = f"{current_record.content}\n{line}"
                continuation_lines += 1
            else:
                unparsed_sheet.append(
                    [
                        source_log_file,
                        line_number,
                        remove_illegal_characters(line),
                        str(exc),
                    ]
                )
                unparsed_lines += 1
            continue

        flush_current_record()
        current_record = next_record

    flush_current_record()

    summary_sheet.append(["项目", "数值"])
    summary_sheet.append(["输入文件", str(input_path)])
    summary_sheet.append(["输出文件", str(output_path)])
    summary_sheet.append(["物理行总数", total_lines])
    summary_sheet.append(["日志记录数", log_records])
    summary_sheet.append(["多行日志续行数", continuation_lines])
    summary_sheet.append(["无法归属行数", unparsed_lines])
    summary_sheet.append([])

    row = 10
    row = write_counter_section(summary_sheet, "日志文件统计", log_file_counter, row)
    row = write_counter_section(summary_sheet, "模块名称统计", module_counter, row)
    row = write_counter_section(summary_sheet, "等级统计", level_counter, row)
    write_counter_section(summary_sheet, "助记符统计", mnemonic_counter, row)

    set_logs_sheet_style(logs_sheet)
    set_summary_sheet_style(summary_sheet)
    set_unparsed_sheet_style(unparsed_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return {
        "input": str(input_path),
        "output": str(output_path),
        "total_lines": total_lines,
        "log_records": log_records,
        "continuation_lines": continuation_lines,
        "unparsed_lines": unparsed_lines,
    }


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="将设备 logfile 原始文件或日志压缩包转换为 Excel 表格。"
    )
    parser.add_argument("--version", action="version", version=f"%(prog)s {__version__}")
    parser.add_argument("input", help="输入文件，支持 .tar.gz/.tgz、.log.gz 或 .log")
    parser.add_argument("-o", "--output", help="输出 Excel 文件名，固定保存到脚本目录的 output 文件夹")
    parser.add_argument("--open", action="store_true", help="完成后在 macOS 中打开输出文件")
    return parser


def main(argv: Optional[list[str]] = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        parser.error(f"输入文件不存在：{input_path}")

    output_path = (
        requested_output_path(args.output)
        if args.output
        else default_output_path(input_path)
    )

    log(f"开始处理：{input_path}")
    result = parse_logfile_to_excel(input_path, output_path)
    log(
        "处理完成："
        f"物理行 {result['total_lines']:,}，"
        f"日志记录 {result['log_records']:,}，"
        f"续行 {result['continuation_lines']:,}，"
        f"无法归属 {result['unparsed_lines']:,}"
    )
    log(f"已生成：{output_path}")

    if args.open:
        subprocess.run(["open", str(output_path)], check=False)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
